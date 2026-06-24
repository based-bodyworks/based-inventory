#!/usr/bin/env python3
"""Daily sales-velocity table from ShipHero -> CSV (+ XLSX).

WHAT THIS PRODUCES
------------------
Daily units sold per single-SKU over the last N days (default 90), broken out
PER SALES CHANNEL, 0-filled for SKU-days with no sales.

- XLSX: an ABOUT sheet plus one tab per channel -- "DAILY VELOCITY - TIKTOK",
  "DAILY VELOCITY - SHOPIFY", "DAILY VELOCITY - AMAZON (FBM)" -- each Date, SKU,
  QTY SOLD[, SKU CODE]. All tabs share the same SKU rows so they line up.
- CSV: long format Date, SKU, Channel, QTY SOLD[, SKU CODE] -- one file holding
  every channel (sum across channels = blended demand).

Amazon is FBM-only: ShipHero sees merchant-fulfilled Amazon orders (shipped from
Merchdrop), NOT Amazon FBA (ships from Amazon's warehouses, never hits this
account), so the Amazon tab undercounts any FBA-fulfilled SKU.

`SKU` is the ShipHero product display name (e.g. "Texture Powder"); the
trailing `SKU CODE` column carries the raw ShipHero SKU so display-name
collisions stay disambiguated (drop it with --no-sku-code). Rows are sorted by
Date ascending, then SKU, then code.

DEFINITION OF "SOLD" (the DEFAULT, recommended path)
----------------------------------------------------
QTY SOLD = units **ORDERED** per day (demand), read from the ShipHero
`orders` endpoint by order_date, with kit/bundle demand resolved to
component SKUs (via inventory.aggregate_orders_demand). ShipHero order
line_items contain BOTH the kit SKU and its components pre-exploded as
separate lines; the aggregator counts the pre-listed component lines and
treats the kit line as a marker, only exploding the kit definition when
the components are NOT pre-listed. This is one efficient pass over orders
and reflects true demand even during out-of-stock periods.

Caveat: bundle explosion applies CURRENT kit recipes to historical orders,
so it is slightly off if a kit's components changed inside the window.
(The alternative "units shipped" path via inventory_changes is exact but
per-SKU and far slower; this script intentionally uses the demand path.)

SCOPE
-----
Rows = single/component SKUs that actually SOLD in the window (default; pass
--all-active-skus to instead 0-fill every active non-kit SKU). Kits/bundles get
NO row of their own; their sales roll INTO the component SKUs. Warehouse =
Merchdrop.

Only real sales channels are counted (shop_name allowlist: BASED / Shopify /
Amazon). Internal "Manual Order" records -- ShipHero's label for manually
created orders and kit-build / component adjustments -- are excluded, because
they are not customer demand (a single Manual Order of 300k packaging units for
a Curl Mousse build was inflating a sample day by 300k). Every excluded channel
is reported in the run summary and the README so nothing is dropped silently;
override the allowlist with --include-shop / --include-all-shops.

REUSE
-----
Auth, the credit-aware client, orders+line_items(first:15) pagination, the
kit/registry builder and the bundle-explosion aggregator all come from the
existing based_inventory package. The only new logic here is bucketing by
day and emitting the dense 0-filled grid (both unit-tested).

RUNTIME / RESUMABILITY
----------------------
A 90-day pull is a long batch (ShipHero's ~60 credits/sec refill paces the
orders calls; expect well over an hour). Progress is checkpointed to disk
per day, so a throttle or crash never loses completed days -- just re-run
the same command and it resumes from the first un-fetched day.

USAGE
-----
    cd ~/tools/based-inventory
    .venv/bin/python scripts/daily_sales_velocity.py            # last 90 days
    .venv/bin/python scripts/daily_sales_velocity.py --days 7   # quick smoke
    .venv/bin/python scripts/daily_sales_velocity.py --start-date 2025-12-03 --end-date 2026-06-03  # ~6 months
    .venv/bin/python scripts/daily_sales_velocity.py --fresh    # ignore checkpoint

Reads SHIPHERO_REFRESH_TOKEN (preferred) / SHIPHERO_ACCESS_TOKEN from .env.
The refresh token is never printed.
"""

from __future__ import annotations

import argparse
import json
import logging
import os
import signal
import sys
import time
from datetime import date, datetime, timedelta
from pathlib import Path

import requests

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "src"))

from dotenv import load_dotenv  # noqa: E402

from based_inventory.inventory import aggregate_orders_demand  # noqa: E402
from based_inventory.registry import BundleRegistry, build_registry  # noqa: E402
from based_inventory.shiphero import (  # noqa: E402
    MERCHDROP_WAREHOUSE_ID,
    ShipHeroClient,
    WarehouseStock,
)
from based_inventory.shiphero_auth import resolve_access_token  # noqa: E402

logger = logging.getLogger("daily_sales_velocity")

DATA_DIR = ROOT / "data"
COMPONENTS_PATH = DATA_DIR / "set-components.json"
DEFAULT_OUTPUT_DIR = DATA_DIR / "sales-velocity"

# Leave headroom for other ShipHero consumers, but lower than the alert
# job's 3400 floor: a long backfill should drain a little deeper so the
# refill wait between calls stays closer to the cost of a single call.
DEFAULT_CREDITS_FLOOR = 2000

# Generous per-day page budget: a single day paginates 100 orders/page, so
# 500 pages covers up to ~50k orders/day -- well beyond any real Based day.
DEFAULT_MAX_PAGES_PER_DAY = 500

DEFINITION = "ordered"  # see module docstring; "shipped" is the unused alternative

# The real DTC/retail sales channels (ShipHero shop_name values). Orders on any
# OTHER channel -- e.g. "Manual Order", ShipHero's label for manually-created
# orders and kit-build / component adjustments -- are NOT customer demand and
# are excluded by default. (A single "Manual Order" of 300k packaging units for
# a Curl Mousse build was inflating one sample day by 300k.) This is a strict
# allowlist: a NEW sales channel is dropped until added here, so every excluded
# channel is logged loudly per the run summary. Override with --include-shop.
DEFAULT_SALES_CHANNELS = (
    "BASED",  # TikTok Shop
    "basedbodyworks.myshopify.com",  # Shopify
    "Based Bodyworks Amazon",  # Amazon
)

# Friendly tab/CSV labels per shop_name, and the order channel tabs appear in.
# Amazon is "(FBM)" because ShipHero only sees merchant-fulfilled Amazon orders
# (shipped from Merchdrop); Amazon FBA ships from Amazon's own warehouses and
# never flows through this account, so the Amazon figures undercount FBA SKUs.
AMAZON_SHOP = "Based Bodyworks Amazon"
CHANNEL_LABELS = {
    AMAZON_SHOP: "AMAZON (FBM)",
    "basedbodyworks.myshopify.com": "SHOPIFY",
    "BASED": "TIKTOK",
}
CHANNEL_ORDER = (
    "Based Bodyworks Amazon",
    "basedbodyworks.myshopify.com",
    "BASED",
)

# Checkpoint schema. v1 stored a single combined {sku: units} per day; v2 stores
# per-channel {shop_name: {sku: units}}. A v1 checkpoint is discarded on load.
CHECKPOINT_SCHEMA = 2


# ---------------------------------------------------------------------------
# Pure helpers (unit-tested in tests/test_daily_sales_velocity.py)
# ---------------------------------------------------------------------------
def date_range(start: date, end: date) -> list[date]:
    """Inclusive list of dates from start to end."""
    if end < start:
        return []
    return [start + timedelta(days=i) for i in range((end - start).days + 1)]


def resolve_start_date(end_date: date, days: int, start_date_str: str | None = None) -> date:
    """Window start: explicit --start-date if given, else end_date - (days-1)."""
    if start_date_str:
        return datetime.strptime(start_date_str, "%Y-%m-%d").date()
    return end_date - timedelta(days=days - 1)


def day_bounds(day_str: str) -> tuple[str, str]:
    """('YYYY-MM-DDT00:00:00', next-day 'YYYY-MM-DDT00:00:00') for a day string."""
    d = datetime.strptime(day_str, "%Y-%m-%d")
    start = d.strftime("%Y-%m-%dT00:00:00")
    end = (d + timedelta(days=1)).strftime("%Y-%m-%dT00:00:00")
    return start, end


def select_sales_orders(
    orders: list[dict], include_shops: set[str] | None
) -> tuple[list[dict], dict[str, int]]:
    """Split orders into (kept sales orders, excluded-shop_name -> order count).

    include_shops is the allowlist of real sales-channel shop_name values; when
    None, no channel filter is applied (every order is kept). Orders whose
    shop_name is not in the allowlist are dropped and counted by shop_name so
    the run can report exactly what it excluded (a new sales channel must not be
    silently lost).
    """
    if include_shops is None:
        return list(orders), {}
    kept: list[dict] = []
    excluded: dict[str, int] = {}
    for o in orders:
        shop = o.get("shop_name") or "(none)"
        if shop in include_shops:
            kept.append(o)
        else:
            excluded[shop] = excluded.get(shop, 0) + 1
    return kept, excluded


def units_ordered_for_day(
    orders: list[dict], registry: BundleRegistry, day_str: str
) -> dict[str, int]:
    """Exploded component-SKU -> units ordered, for orders whose order_date is day_str.

    Filters to the target day by order_date[:10] (idempotent regardless of
    whether ShipHero's order_date_to bound is inclusive, so an order landing
    on a midnight boundary is attributed to exactly one day). Kit/bundle
    demand is resolved to component SKUs by aggregate_orders_demand (pre-
    exploded component lines counted once; kit line as marker).

    Channel filtering is applied UPSTREAM (see select_sales_orders); this
    function assumes `orders` is already restricted to real sales channels.
    """
    day_orders = [o for o in orders if (o.get("order_date") or "")[:10] == day_str]
    demand = aggregate_orders_demand(day_orders, registry)
    return {sku: sample.units for sku, sample in demand.items() if sample.units}


def units_ordered_by_channel(
    orders: list[dict], registry: BundleRegistry, day_str: str, channels: set[str]
) -> dict[str, dict[str, int]]:
    """Per-channel {shop_name -> {sku -> units}} for one day.

    Splits `orders` by shop_name and runs the same day-filter + bundle explosion
    per channel, so each channel's velocity is tracked independently. Every shop
    in `channels` gets an entry (empty if it had no sales that day) so the tabs
    stay stable across the window.
    """
    out: dict[str, dict[str, int]] = {}
    for shop in channels:
        shop_orders = [o for o in orders if (o.get("shop_name") or "(none)") == shop]
        out[shop] = units_ordered_for_day(shop_orders, registry, day_str)
    return out


def ordered_channels(channel_keys: set[str]) -> list[str]:
    """Channels in display order: known channels (CHANNEL_ORDER) first, then the rest."""
    keys = set(channel_keys)
    ordered = [s for s in CHANNEL_ORDER if s in keys]
    ordered += sorted(k for k in keys if k not in CHANNEL_ORDER)
    return ordered


def channel_label(shop_name: str) -> str:
    """Friendly label for a shop_name (falls back to the raw name)."""
    return CHANNEL_LABELS.get(shop_name, shop_name)


def _sheet_title(label: str) -> str:
    """Excel-safe sheet title (<=31 chars, no \\ / ? * [ ] :)."""
    title = f"DAILY VELOCITY - {label}"
    for bad in "\\/?*[]:":
        title = title.replace(bad, "-")
    return title[:31]


def build_sku_universe(
    stock: list[WarehouseStock],
    registry: BundleRegistry,
    demand_skus: set[str],
    sold_only: bool = True,
) -> tuple[set[str], dict[str, str]]:
    """Return (SKU set to give rows, sku->display-name map).

    sold_only=True (default): rows = exactly the SKUs that received real demand
    in the window (kits/bundles excluded -- their demand already rolled into
    components). This drops the long tail of never-selling packaging / raw-
    material / test SKUs and matches the "things that sell" shape.

    sold_only=False: rows = every active non-kit, non-bundle SKU (UNIONed with
    any demand-only SKU), i.e. the full active catalogue 0-filled even where a
    SKU never sold.
    """
    kit_skus = {s.sku for s in stock if s.is_kit}
    name_map: dict[str, str] = {s.sku: (s.product_name or s.sku) for s in stock}
    demand_singles = {d for d in demand_skus if d not in registry.bundle_skus and d not in kit_skus}
    if sold_only:
        skus = demand_singles
    else:
        base = {s.sku for s in stock if not s.is_kit and s.sku not in registry.bundle_skus}
        skus = base | demand_singles
    for d in skus:
        name_map.setdefault(d, d)
    return skus, name_map


def build_grid_rows(
    day_maps: dict[str, dict[str, int]],
    sku_codes: set[str],
    name_map: dict[str, str],
    day_strs: list[str],
) -> list[tuple[str, str, int, str]]:
    """Dense, 0-filled grid: one (date, name, qty, code) tuple per day x SKU.

    Sorted by date ascending, then display name (case-insensitive), then code.
    """
    codes = sorted(sku_codes)
    rows: list[tuple[str, str, int, str]] = []
    for day in day_strs:
        dm = day_maps.get(day, {})
        for code in codes:
            qty = int(dm.get(code, 0) or 0)
            rows.append((day, name_map.get(code, code), qty, code))
    rows.sort(key=lambda r: (r[0], r[1].lower(), r[3]))
    return rows


# ---------------------------------------------------------------------------
# Output writers
# ---------------------------------------------------------------------------
def write_csv(
    channel_rows: list[tuple[str, list[tuple[str, str, int, str]]]],
    path: Path,
    include_code: bool = True,
) -> None:
    """Long-format CSV: Date, SKU, Channel, QTY SOLD[, SKU CODE].

    channel_rows is [(channel_label, rows), ...]; one output row per
    date x SKU x channel (0-filled), so the file holds every channel and sums to
    the blended total. Sorted by date, then SKU, then channel.
    """
    import csv

    flat: list[tuple] = []
    for label, rows in channel_rows:
        for day, name, qty, code in rows:
            flat.append((day, name, label, qty, code))
    flat.sort(key=lambda r: (r[0], r[1].lower(), r[2], r[4]))

    with open(path, "w", newline="") as f:
        w = csv.writer(f)
        header = ["Date", "SKU", "Channel", "QTY SOLD"]
        if include_code:
            header.append("SKU CODE")
        w.writerow(header)
        for day, name, label, qty, code in flat:
            row = [day, name, label, qty]
            if include_code:
                row.append(code)
            w.writerow(row)


def write_xlsx(
    sheets: list[tuple[str, list[tuple[str, str, int, str]]]],
    path: Path,
    about_lines: list[str],
    include_code: bool = True,
) -> None:
    """ABOUT sheet + one tab per (title, rows) in `sheets`."""
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    about = wb.active
    about.title = "ABOUT"
    for i, line in enumerate(about_lines, start=1):
        about.cell(row=i, column=1, value=line)
    about.column_dimensions["A"].width = 110

    header = ["Date", "SKU", "QTY SOLD"] + (["SKU CODE"] if include_code else [])
    widths = [12, 34, 10, 26]
    for title, rows in sheets:
        ws = wb.create_sheet(title)
        ws.append(header)
        for c in range(1, len(header) + 1):
            ws.cell(row=1, column=c).font = Font(bold=True)
        for day, name, qty, code in rows:
            ws.append([day, name, qty] + ([code] if include_code else []))
        ws.freeze_panes = "A2"
        for i, wdt in enumerate(widths[: len(header)], start=1):
            ws.column_dimensions[get_column_letter(i)].width = wdt

    wb.save(path)


# ---------------------------------------------------------------------------
# Checkpoint
# ---------------------------------------------------------------------------
def load_checkpoint(path: Path) -> dict | None:
    if not path.exists():
        return None
    try:
        return json.loads(path.read_text())
    except (json.JSONDecodeError, OSError):
        logger.warning("Checkpoint at %s is unreadable; starting fresh.", path)
        return None


def save_checkpoint(path: Path, state: dict) -> None:
    tmp = path.with_suffix(path.suffix + ".tmp")
    tmp.write_text(json.dumps(state))
    tmp.replace(path)  # atomic on POSIX


# ---------------------------------------------------------------------------
# Orchestration
# ---------------------------------------------------------------------------
def _build_client(credits_floor: int) -> ShipHeroClient:
    refresh = os.environ.get("SHIPHERO_REFRESH_TOKEN")
    fallback = os.environ.get("SHIPHERO_ACCESS_TOKEN")
    if not refresh and not fallback:
        sys.exit("ERROR: neither SHIPHERO_REFRESH_TOKEN nor SHIPHERO_ACCESS_TOKEN in .env")
    token = resolve_access_token(refresh_token=refresh, fallback_access_token=fallback)
    api_url = os.environ.get("SHIPHERO_API_URL") or "https://public-api.shiphero.com/graphql"
    return ShipHeroClient(token=token, api_url=api_url, min_credits_floor=credits_floor)


def _load_catalog(client: ShipHeroClient) -> tuple[list[WarehouseStock], BundleRegistry]:
    """Active warehouse stock + the merged bundle registry (with component backfill)."""
    logger.info("Loading warehouse stock (Merchdrop)...")
    stock = client.fetch_warehouse_stock(warehouse_id=MERCHDROP_WAREHOUSE_ID)
    logger.info("  %d active warehouse_products", len(stock))

    logger.info("Loading kit definitions...")
    kits = client.fetch_all_kits()
    logger.info("  %d kits", len(kits))

    # Backfill any component SKU missing from the warehouse page (so the
    # registry + display names see it). One targeted lookup each.
    component_skus = {c[0] for k in kits for c in k.components}
    known = {s.sku for s in stock}
    missing = sorted(component_skus - known)
    if missing:
        logger.info("  backfilling %d component SKUs missing from page-1 stock", len(missing))
    for sku in missing:
        try:
            row = client.fetch_warehouse_product_for_sku(sku, MERCHDROP_WAREHOUSE_ID)
            if row is not None:
                stock.append(row)
        except RuntimeError:
            continue

    registry = build_registry(kits, stock, COMPONENTS_PATH)
    logger.info("  %d bundle SKUs in registry", len(registry.bundle_skus))
    return stock, registry


# Network errors that should pause-and-retry the day rather than kill the run.
_RETRYABLE_NETWORK_HINTS = (
    "network error",
    "failed to resolve",
    "max retries exceeded",
    "connection",
    "timed out",
    "timeout",
    "name resolution",
    "temporarily",
    "bad gateway",
    "gateway time",
    "503",
    "502",
)
# Generous cap so a long outage (overnight sleep without caffeinate, ISP drop)
# is ridden through; progress is checkpointed per day regardless.
_DAY_FETCH_MAX_NETWORK_RETRIES = 240

# Hard wall-clock cap on a SINGLE day's fetch. A healthy day finishes in well
# under this even at peak volume (~8-16 min); anything past it is a wedged
# socket (a reset connection that never times out on read), so we abort and
# retry the day. Without this, a hung read can stall the whole run silently.
_DAY_FETCH_WATCHDOG_SECS = 1500


class _FetchTimeoutError(Exception):
    """Raised by the per-day watchdog when a single day's fetch hangs."""


def _watchdog_handler(signum, frame):
    raise _FetchTimeoutError("per-day fetch watchdog fired")


def _set_watchdog(secs: int):
    """Arm a SIGALRM watchdog; returns the previous handler (or None if unavailable)."""
    try:
        previous = signal.signal(signal.SIGALRM, _watchdog_handler)
        signal.alarm(secs)
        return previous
    except (ValueError, OSError):  # not the main thread / unsupported platform
        return None


def _clear_watchdog(previous) -> None:
    if previous is None:
        return
    try:
        signal.alarm(0)
        signal.signal(signal.SIGALRM, previous)
    except (ValueError, OSError):
        pass


def _is_retryable_network_error(exc: Exception) -> bool:
    """True for transient connectivity failures (drops, DNS, 5xx, hangs) worth retrying."""
    if isinstance(exc, requests.RequestException):
        return True
    if isinstance(exc, _FetchTimeoutError):
        return True
    if isinstance(exc, RuntimeError):
        msg = str(exc).lower()
        return any(h in msg for h in _RETRYABLE_NETWORK_HINTS)
    return False


def _fetch_orders_resilient(
    client: ShipHeroClient, day: str, since: str, until: str, max_pages: int
) -> list[dict]:
    """fetch_orders_window that PAUSES and retries through network drops / sleep / hangs.

    A transient connectivity failure (wifi blip, DNS failure, sleep/wake, a
    ShipHero 5xx) OR a wedged socket that hangs past the per-day watchdog waits
    with escalating backoff and retries the SAME day instead of crashing or
    stalling the whole run. Non-network errors propagate. Either way progress is
    checkpointed per completed day, so nothing already pulled is lost.
    """
    attempt = 0
    while True:
        previous = _set_watchdog(_DAY_FETCH_WATCHDOG_SECS)
        try:
            orders = client.fetch_orders_window(
                since_iso=since,
                warehouse_id=MERCHDROP_WAREHOUSE_ID,
                max_pages=max_pages,
                until_iso=until,
            )
        except (RuntimeError, requests.RequestException, _FetchTimeoutError) as e:
            _clear_watchdog(previous)  # disarm BEFORE the backoff sleep
            if not _is_retryable_network_error(e):
                raise
            attempt += 1
            if attempt > _DAY_FETCH_MAX_NETWORK_RETRIES:
                raise RuntimeError(
                    f"Network unavailable after {attempt} retries on {day}; progress is "
                    "checkpointed -- re-run the same command to resume."
                ) from e
            wait = min(30 * (2 ** min(attempt - 1, 4)), 300)  # 30,60,120,240,300,...
            logger.warning(
                "Network issue on %s (retry %d): %s -- waiting %ds then retrying.",
                day,
                attempt,
                str(e)[:120],
                wait,
            )
            time.sleep(wait)
        else:
            _clear_watchdog(previous)  # disarm on success before returning to caller
            return orders


def run(args: argparse.Namespace) -> int:
    out_dir = Path(args.out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    end_date = (
        datetime.strptime(args.end_date, "%Y-%m-%d").date()
        if args.end_date
        else datetime.utcnow().date()
    )
    start_date = resolve_start_date(end_date, args.days, args.start_date)
    day_strs = [d.strftime("%Y-%m-%d") for d in date_range(start_date, end_date)]
    slug = f"daily-sales-velocity_{start_date.isoformat()}_{end_date.isoformat()}"
    ckpt_path = out_dir / f"{slug}.checkpoint.json"

    include_shops = None if args.include_all_shops else set(args.include_shop)
    include_shops_key = sorted(include_shops) if include_shops is not None else None

    logger.info(
        "Window: %s .. %s (%d days, UTC). QTY SOLD = units %s/day.",
        start_date,
        end_date,
        len(day_strs),
        DEFINITION,
    )
    logger.info(
        "Sales channels counted: %s",
        ", ".join(include_shops_key) if include_shops_key else "ALL (no channel filter)",
    )

    client = _build_client(args.credits_floor)
    stock, registry = _load_catalog(client)

    # --- checkpoint state ---
    # The cached per-day, per-channel maps are channel-filtered + bundle-
    # exploded, so a change to the definition, warehouse, channel allowlist, or
    # checkpoint schema invalidates them.
    state = None if args.fresh else load_checkpoint(ckpt_path)
    if state and (
        state.get("schema") != CHECKPOINT_SCHEMA
        or state.get("definition") != DEFINITION
        or state.get("warehouse_id") != MERCHDROP_WAREHOUSE_ID
        or state.get("include_shops") != include_shops_key
    ):
        logger.warning("Checkpoint schema/definition/warehouse/channel mismatch; starting fresh.")
        state = None
    if state is None:
        state = {
            "schema": CHECKPOINT_SCHEMA,
            "definition": DEFINITION,
            "warehouse_id": MERCHDROP_WAREHOUSE_ID,
            "line_items_cap": ShipHeroClient._ORDERS_LINE_ITEMS_CAP,
            "include_shops": include_shops_key,
            "days": {},
            "excluded_shops": {},
        }
    # days_done[day] = {shop_name: {sku: units}}  (per-channel)
    days_done: dict[str, dict[str, dict[str, int]]] = state["days"]
    excluded_shops: dict[str, int] = state.setdefault("excluded_shops", {})

    # --- fetch each day not already checkpointed ---
    todo = [d for d in day_strs if d not in days_done]
    logger.info(
        "%d/%d days already cached; fetching %d.",
        len(day_strs) - len(todo),
        len(day_strs),
        len(todo),
    )
    for idx, day in enumerate(todo, start=1):
        t0 = time.time()
        since, until = day_bounds(day)
        orders = _fetch_orders_resilient(client, day, since, until, args.max_pages)
        sales_orders, excluded = select_sales_orders(orders, include_shops)
        for shop, n in excluded.items():
            excluded_shops[shop] = excluded_shops.get(shop, 0) + n
        # Which channels to break out: the allowlist when set, else whatever
        # shop_names actually appear in the kept orders.
        day_channels = (
            include_shops
            if include_shops is not None
            else {o.get("shop_name") or "(none)" for o in sales_orders}
        )
        channel_maps = units_ordered_by_channel(sales_orders, registry, day, day_channels)
        days_done[day] = channel_maps
        save_checkpoint(ckpt_path, state)
        per_ch = ", ".join(
            f"{channel_label(s)}:{sum(channel_maps[s].values())}"
            for s in ordered_channels(set(channel_maps))
        )
        excl_note = ""
        if excluded:
            excl_note = " | EXCLUDED " + ", ".join(
                f"{shop}:{n}" for shop, n in sorted(excluded.items(), key=lambda x: -x[1])
            )
        logger.info(
            "[%d/%d] %s: %d orders (%d sales) -> [%s] (%.1fs)%s",
            idx,
            len(todo),
            day,
            len(orders),
            len(sales_orders),
            per_ch,
            time.time() - t0,
            excl_note,
        )

    # --- build a dense grid PER CHANNEL over a shared SKU universe ---
    # Channels to emit (a tab each), and the shared row universe = every SKU that
    # sold on ANY channel in the window, so all channel tabs line up row-for-row.
    all_channels: set[str] = set()
    for day in day_strs:
        all_channels |= set(days_done.get(day, {}).keys())
    if include_shops is not None:
        all_channels |= set(include_shops)  # guarantee a tab for every allowlisted channel
    emit_channels = ordered_channels(all_channels)

    demand_skus: set[str] = set()
    for day in day_strs:
        for ch_map in days_done.get(day, {}).values():
            demand_skus |= set(ch_map.keys())
    sku_codes, name_map = build_sku_universe(
        stock, registry, demand_skus, sold_only=not args.all_active_skus
    )

    # Per channel: a {day: {sku: units}} view, then the dense 0-filled grid.
    channel_rows: list[tuple[str, list[tuple[str, str, int, str]]]] = []
    channel_totals: list[tuple[str, int]] = []
    for shop in emit_channels:
        ch_day_maps = {day: days_done.get(day, {}).get(shop, {}) for day in day_strs}
        rows = build_grid_rows(ch_day_maps, sku_codes, name_map, day_strs)
        label = channel_label(shop)
        channel_rows.append((label, rows))
        channel_totals.append((label, sum(r[2] for r in rows)))

    total_units = sum(t for _label, t in channel_totals)
    channels_line = (
        "Sales channels counted (shop_name allowlist): " + ", ".join(include_shops_key)
        if include_shops_key
        else "Sales channels: ALL (no channel filter applied)"
    )
    excluded_line = (
        "Excluded non-sales channels (orders dropped): "
        + ", ".join(
            f"{shop} ({n})" for shop, n in sorted(excluded_shops.items(), key=lambda x: -x[1])
        )
        if excluded_shops
        else "Excluded non-sales channels: none seen in window."
    )
    scope_line = (
        "Scope: every active non-kit SKU (0-filled even where it never sold); kits/bundles get no row."
        if args.all_active_skus
        else "Scope: only SKUs that sold in-window (kits/bundles excluded; their sales rolled into components)."
    )
    per_channel_line = "Per-channel total units: " + " | ".join(
        f"{label}: {t}" for label, t in channel_totals
    )
    about_lines = [
        "Based BodyWorks - Daily Sales Velocity (ShipHero)",
        "QTY SOLD = units ORDERED per day (demand), from ShipHero `orders` by order_date.",
        "Kit/bundle demand is resolved to component SKUs: ShipHero pre-explodes kit components",
        "as their own lines, so those lines are counted and the kit line is a marker; a kit",
        "whose components are NOT pre-listed is exploded using CURRENT kit recipes",
        "(slightly off if a kit's components changed inside the window).",
        "One tab per sales channel (TikTok / Shopify / Amazon); same SKU rows on each tab.",
        f"Warehouse: Merchdrop ({MERCHDROP_WAREHOUSE_ID}).",
        channels_line,
        excluded_line,
        f"Nested line_items capped at first:{ShipHeroClient._ORDERS_LINE_ITEMS_CAP} per order (rare >15-line orders truncate their tail).",
        f"Window: {start_date} .. {end_date} ({len(day_strs)} days, UTC). The most recent day may be partial.",
        scope_line,
        f"SKUs (rows per tab): {len(sku_codes)}  |  Total units ordered (all channels): {total_units}",
        per_channel_line,
        f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')} UTC",
    ]
    if AMAZON_SHOP in emit_channels:
        about_lines.insert(
            6,
            "AMAZON = FBM ONLY (Merchdrop-fulfilled). Amazon FBA ships from Amazon's "
            "warehouses and is NOT in ShipHero, so Amazon undercounts any FBA-fulfilled SKU.",
        )
    if include_shops_key and excluded_shops:
        about_lines.insert(
            9 if AMAZON_SHOP in emit_channels else 8,
            "NOTE: strict allowlist -- if any excluded channel above is a real sales channel, "
            "re-run with --include-shop '<name>'.",
        )

    # CSV is long-format (a Channel column) so it carries every channel in one
    # file; the XLSX gives each channel its own tab.
    csv_path = out_dir / f"{slug}.csv"
    write_csv(channel_rows, csv_path, include_code=not args.no_sku_code)
    readme_path = out_dir / f"{slug}.README.txt"
    readme_path.write_text("\n".join(about_lines) + "\n")
    written = [csv_path, readme_path]

    if not args.no_xlsx:
        try:
            xlsx_path = out_dir / f"{slug}.xlsx"
            sheets = [(_sheet_title(label), rows) for label, rows in channel_rows]
            write_xlsx(sheets, xlsx_path, about_lines, include_code=not args.no_sku_code)
            written.append(xlsx_path)
        except ImportError:
            logger.warning("openpyxl not available; skipped XLSX (CSV written).")

    print("\n".join(about_lines))
    print("\nWrote:")
    for p in written:
        print(f"  {p}")
    print(f"  {ckpt_path}  (checkpoint; safe to delete once you have the CSV)")
    return 0


def build_arg_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(description="Daily ShipHero sales-velocity table -> CSV/XLSX.")
    p.add_argument(
        "--days",
        type=int,
        default=90,
        help="Trailing window length when --start-date is omitted (default 90).",
    )
    p.add_argument(
        "--start-date",
        default=None,
        help="First day of the window, YYYY-MM-DD (overrides --days). E.g. 2025-12-03 for ~6 months.",
    )
    p.add_argument(
        "--end-date",
        default=None,
        help="Last day of the window, YYYY-MM-DD (default: today UTC, a partial day).",
    )
    p.add_argument(
        "--out-dir",
        default=str(DEFAULT_OUTPUT_DIR),
        help=f"Output directory (default: {DEFAULT_OUTPUT_DIR}).",
    )
    p.add_argument(
        "--credits-floor",
        type=int,
        default=DEFAULT_CREDITS_FLOOR,
        help=f"Keep the ShipHero credit pool above this (default {DEFAULT_CREDITS_FLOOR}).",
    )
    p.add_argument(
        "--max-pages",
        type=int,
        default=DEFAULT_MAX_PAGES_PER_DAY,
        help=f"Max order pages per day (default {DEFAULT_MAX_PAGES_PER_DAY}).",
    )
    p.add_argument(
        "--include-shop",
        action="append",
        default=list(DEFAULT_SALES_CHANNELS),
        metavar="SHOP_NAME",
        help="Sales-channel shop_name to count (repeatable). Defaults to the 3 real "
        "channels: BASED, basedbodyworks.myshopify.com, Based Bodyworks Amazon.",
    )
    p.add_argument(
        "--include-all-shops",
        action="store_true",
        help="Disable the channel allowlist and count every order (incl. Manual Order).",
    )
    p.add_argument(
        "--all-active-skus",
        action="store_true",
        help="Give a row to every active non-kit SKU (0-filled), not just SKUs that sold.",
    )
    p.add_argument("--no-xlsx", action="store_true", help="Skip the XLSX output.")
    p.add_argument(
        "--no-sku-code",
        action="store_true",
        help="Drop the trailing SKU CODE column (exact 3-column Date,SKU,QTY SOLD shape).",
    )
    p.add_argument(
        "--fresh", action="store_true", help="Ignore any existing checkpoint and refetch."
    )
    p.add_argument("--log-level", default="INFO", help="Logging level (default INFO).")
    return p


def main(argv: list[str] | None = None) -> int:
    args = build_arg_parser().parse_args(argv)
    # Load .env at run time (NOT import time) so importing this module for tests
    # doesn't leak real env into the pytest session.
    load_dotenv(ROOT / ".env")
    logging.basicConfig(
        level=getattr(logging, args.log_level.upper(), logging.INFO),
        format="%(asctime)s %(levelname)s %(name)s: %(message)s",
    )
    return run(args)


if __name__ == "__main__":
    raise SystemExit(main())
