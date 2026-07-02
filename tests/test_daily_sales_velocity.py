"""Tests for the daily sales-velocity script's new logic.

Covers the genuinely-new pieces (the rest is reused, already-tested package
code): per-day bucketing + bundle explosion, the dense 0-filled grid, the
single-SKU universe selection, and the additive `until_iso` upper bound on
ShipHeroClient.fetch_orders_window.
"""

from __future__ import annotations

import sys
from datetime import date
from pathlib import Path

SCRIPTS = Path(__file__).resolve().parents[1] / "scripts"
sys.path.insert(0, str(SCRIPTS))

import daily_sales_velocity as dsv  # noqa: E402

from based_inventory.registry import BundleEntry, BundleRegistry  # noqa: E402
from based_inventory.shiphero import ShipHeroClient, WarehouseStock  # noqa: E402


# --- fixtures ---------------------------------------------------------------
def _stock(sku: str, name: str, is_kit: bool = False) -> WarehouseStock:
    return WarehouseStock(
        sku=sku,
        on_hand=0,
        available=0,
        allocated=0,
        backorder=0,
        reserve_inventory=0,
        sell_ahead=0,
        product_name=name,
        is_kit=is_kit,
    )


def _order(order_date: str, *line_items: tuple[str, int]) -> dict:
    return {
        "order_number": order_date + "-" + str(hash(line_items) % 10_000),
        "order_date": order_date,
        "line_items": {"edges": [{"node": {"sku": s, "quantity": q}} for s, q in line_items]},
    }


def _registry() -> BundleRegistry:
    """KIT1 = 1x BB-CC (Curl Cream) + 2x BB-SHMP (Shampoo)."""
    kit = BundleEntry(
        bundle_sku="KIT1",
        bundle_name="Curly Kit",
        source="shiphero",
        components_resolved=(("BB-CC", "Curl Cream", 1), ("BB-SHMP", "Shampoo", 2)),
        partially_resolved=False,
    )
    return BundleRegistry(
        bundles=(kit,), bundle_skus=frozenset({"KIT1"}), by_bundle_sku={"KIT1": kit}
    )


# --- date helpers -----------------------------------------------------------
def test_date_range_inclusive() -> None:
    out = dsv.date_range(date(2026, 3, 1), date(2026, 3, 3))
    assert out == [date(2026, 3, 1), date(2026, 3, 2), date(2026, 3, 3)]


def test_date_range_empty_when_reversed() -> None:
    assert dsv.date_range(date(2026, 3, 3), date(2026, 3, 1)) == []


def test_resolve_start_date_from_days() -> None:
    assert dsv.resolve_start_date(date(2026, 6, 3), 90) == date(2026, 3, 6)


def test_resolve_start_date_explicit_overrides_days() -> None:
    # ~6 months: explicit start wins over --days.
    assert dsv.resolve_start_date(date(2026, 6, 3), 90, "2025-12-03") == date(2025, 12, 3)


def test_day_bounds() -> None:
    assert dsv.day_bounds("2026-03-01") == (
        "2026-03-01T00:00:00",
        "2026-03-02T00:00:00",
    )


# --- per-day bucketing + bundle explosion -----------------------------------
def test_units_ordered_for_day_filters_to_the_day() -> None:
    orders = [
        _order("2026-03-01T10:00:00", ("BB-CC", 3)),
        _order("2026-03-02T09:00:00", ("BB-CC", 100)),  # different day, must be excluded
    ]
    out = dsv.units_ordered_for_day(orders, _registry(), "2026-03-01")
    assert out == {"BB-CC": 3}


def test_units_ordered_for_day_explodes_bundles() -> None:
    # 2x KIT1 on the day -> 2 Curl Cream + 4 Shampoo, plus a standalone Curl Cream.
    orders = [
        _order("2026-03-01T08:00:00", ("KIT1", 2)),
        _order("2026-03-01T12:00:00", ("BB-CC", 1)),
    ]
    out = dsv.units_ordered_for_day(orders, _registry(), "2026-03-01")
    assert out == {"BB-CC": 3, "BB-SHMP": 4}
    assert "KIT1" not in out  # the bundle itself never gets counted as a SKU


def test_units_ordered_for_day_empty_day() -> None:
    assert dsv.units_ordered_for_day([], _registry(), "2026-03-01") == {}


# --- channel allowlist ------------------------------------------------------
def test_select_sales_orders_partitions_and_counts_excluded() -> None:
    orders = [
        {"shop_name": "BASED", "line_items": {"edges": []}},
        {"shop_name": "Manual Order", "line_items": {"edges": []}},
        {"shop_name": "Manual Order", "line_items": {"edges": []}},
        {"shop_name": None, "line_items": {"edges": []}},
    ]
    kept, excluded = dsv.select_sales_orders(orders, include_shops={"BASED"})
    assert len(kept) == 1 and kept[0]["shop_name"] == "BASED"
    assert excluded == {"Manual Order": 2, "(none)": 1}


def test_select_sales_orders_none_allowlist_keeps_all() -> None:
    orders = [{"shop_name": "Manual Order", "line_items": {"edges": []}}]
    kept, excluded = dsv.select_sales_orders(orders, include_shops=None)
    assert len(kept) == 1 and excluded == {}


# --- single-SKU universe ----------------------------------------------------
def test_build_sku_universe_sold_only_is_just_demand_singles() -> None:
    # Default sold_only: rows = SKUs that received demand, kits/bundles excluded.
    stock = [
        _stock("BB-CC", "Curl Cream"),
        _stock("BB-SHMP", "Shampoo"),  # active but no demand -> NO row
        _stock("KIT1", "Curly Kit", is_kit=True),
    ]
    skus, names = dsv.build_sku_universe(stock, _registry(), demand_skus={"BB-CC", "KIT1"})
    assert skus == {"BB-CC"}  # BB-SHMP dropped (no sales); KIT1 dropped (kit)
    assert names["BB-CC"] == "Curl Cream"


def test_build_sku_universe_sold_only_includes_demand_only_component() -> None:
    # A component that sold but isn't on the active warehouse page still gets a row.
    stock = [_stock("BB-CC", "Curl Cream")]
    skus, names = dsv.build_sku_universe(stock, _registry(), demand_skus={"BB-CC", "BB-OLD"})
    assert skus == {"BB-CC", "BB-OLD"}
    assert names["BB-OLD"] == "BB-OLD"  # falls back to the code when no name is known


def test_build_sku_universe_all_active_includes_zero_sale_skus() -> None:
    stock = [
        _stock("BB-CC", "Curl Cream"),
        _stock("BB-SHMP", "Shampoo"),  # no demand
        _stock("KIT1", "Curly Kit", is_kit=True),
    ]
    skus, _ = dsv.build_sku_universe(stock, _registry(), demand_skus={"BB-CC"}, sold_only=False)
    assert skus == {"BB-CC", "BB-SHMP"}  # both active singles; kit still excluded


# --- dense grid -------------------------------------------------------------
def test_build_grid_rows_zero_fills_every_day_and_sku() -> None:
    day_maps = {
        "2026-03-01": {"BB-CC": 3},
        "2026-03-02": {"BB-CC": 5, "BB-SHMP": 2},
    }
    names = {"BB-CC": "Curl Cream", "BB-SHMP": "Shampoo"}
    rows = dsv.build_grid_rows(day_maps, {"BB-CC", "BB-SHMP"}, names, ["2026-03-01", "2026-03-02"])
    # 2 days x 2 SKUs = 4 rows, every cell present (0-filled where no sale).
    assert len(rows) == 4
    as_dict = {(d, code): qty for d, _name, qty, code in rows}
    assert as_dict[("2026-03-01", "BB-CC")] == 3
    assert as_dict[("2026-03-01", "BB-SHMP")] == 0  # 0-filled
    assert as_dict[("2026-03-02", "BB-CC")] == 5
    assert as_dict[("2026-03-02", "BB-SHMP")] == 2


def test_build_grid_rows_sorted_by_date_then_name() -> None:
    day_maps = {"2026-03-02": {}, "2026-03-01": {}}
    names = {"Z-SKU": "Apple", "A-SKU": "Zebra"}
    rows = dsv.build_grid_rows(day_maps, {"A-SKU", "Z-SKU"}, names, ["2026-03-01", "2026-03-02"])
    # date asc first, then name (case-insensitive) -> Apple before Zebra.
    assert [(r[0], r[1]) for r in rows] == [
        ("2026-03-01", "Apple"),
        ("2026-03-01", "Zebra"),
        ("2026-03-02", "Apple"),
        ("2026-03-02", "Zebra"),
    ]


# --- per-channel breakout ---------------------------------------------------
def test_units_ordered_by_channel_partitions_and_explodes() -> None:
    orders = [
        {
            "shop_name": "BASED",
            "order_date": "2026-03-01T01:00:00",
            "line_items": {"edges": [{"node": {"sku": "BB-CC", "quantity": 2}}]},
        },
        {
            "shop_name": "Based Bodyworks Amazon",
            "order_date": "2026-03-01T02:00:00",
            "line_items": {"edges": [{"node": {"sku": "KIT1", "quantity": 1}}]},
        },
    ]
    out = dsv.units_ordered_by_channel(
        orders, _registry(), "2026-03-01", {"BASED", "Based Bodyworks Amazon"}
    )
    assert out["BASED"] == {"BB-CC": 2}
    # KIT1 explodes -> 1 Curl Cream + 2 Shampoo, attributed to Amazon only.
    assert out["Based Bodyworks Amazon"] == {"BB-CC": 1, "BB-SHMP": 2}


def test_units_ordered_by_channel_empty_channel_present() -> None:
    orders = [
        {
            "shop_name": "BASED",
            "order_date": "2026-03-01T01:00:00",
            "line_items": {"edges": [{"node": {"sku": "BB-CC", "quantity": 1}}]},
        },
    ]
    out = dsv.units_ordered_by_channel(orders, _registry(), "2026-03-01", {"BASED", "X"})
    assert out["X"] == {}  # channel with no sales still appears (stable tabs)


def test_ordered_channels_known_first_then_alpha() -> None:
    out = dsv.ordered_channels({"BASED", "Zzz", "Based Bodyworks Amazon", "Aaa"})
    # CHANNEL_ORDER (Amazon, Shopify, TikTok) first for known ones, then others A-Z.
    assert out == ["Based Bodyworks Amazon", "BASED", "Aaa", "Zzz"]


def test_channel_label_maps_known_and_falls_back() -> None:
    assert dsv.channel_label("BASED") == "TIKTOK"
    assert dsv.channel_label("Based Bodyworks Amazon") == "AMAZON (FBM)"  # FBA not in ShipHero
    assert dsv.channel_label("Some New Channel") == "Some New Channel"


# --- raw (un-exploded) bundle sales view ------------------------------------
def test_bundle_units_by_channel_counts_raw_not_exploded() -> None:
    # 2x KIT1 as-sold stays 2x KIT1 (no component explosion); the standalone
    # single BB-CC is NOT a bundle so it never appears in this view.
    orders = [
        {
            "shop_name": "BASED",
            "order_date": "2026-03-01T01:00:00",
            "line_items": {
                "edges": [
                    {"node": {"sku": "KIT1", "quantity": 2}},
                    {"node": {"sku": "BB-CC", "quantity": 1}},
                ]
            },
        },
    ]
    out = dsv.bundle_units_by_channel(orders, _registry(), "2026-03-01", {"BASED"})
    assert out == {"BASED": {"KIT1": 2}}


def test_bundle_units_by_channel_partitions_by_shop_and_keeps_empty_channels() -> None:
    orders = [
        {
            "shop_name": "BASED",
            "order_date": "2026-03-01T01:00:00",
            "line_items": {"edges": [{"node": {"sku": "KIT1", "quantity": 1}}]},
        },
        {
            "shop_name": "Based Bodyworks Amazon",
            "order_date": "2026-03-01T02:00:00",
            "line_items": {"edges": [{"node": {"sku": "KIT1", "quantity": 3}}]},
        },
    ]
    out = dsv.bundle_units_by_channel(
        orders, _registry(), "2026-03-01", {"BASED", "Based Bodyworks Amazon", "X"}
    )
    assert out["BASED"] == {"KIT1": 1}
    assert out["Based Bodyworks Amazon"] == {"KIT1": 3}
    assert out["X"] == {}  # channel with no bundle sales still appears (stable tabs)


def test_bundle_units_by_channel_filters_to_day() -> None:
    orders = [
        {
            "shop_name": "BASED",
            "order_date": "2026-03-02T01:00:00",  # different day, must be excluded
            "line_items": {"edges": [{"node": {"sku": "KIT1", "quantity": 5}}]},
        },
    ]
    out = dsv.bundle_units_by_channel(orders, _registry(), "2026-03-01", {"BASED"})
    assert out == {"BASED": {}}


# --- coverage guards (unregistered bundles must not vanish silently) ---------
def test_unregistered_kit_skus_flags_kits_missing_from_registry() -> None:
    stock = [
        _stock("KIT1", "Curly Kit", is_kit=True),  # registered -> fine
        _stock("NEWKIT", "Brand New Kit", is_kit=True),  # NOT in registry -> flagged
        _stock("BB-CC", "Curl Cream"),  # single -> ignored
    ]
    assert dsv.unregistered_kit_skus(stock, _registry()) == {"NEWKIT"}


def test_suspicious_unexploded_skus_flags_bundleish_codes_and_names() -> None:
    codes = {"BB-CC", "BB-XYZ-2PCK", "MYSTERY1", "BB-SS-2PK", "SET1", "MYSTERY2"}
    names = {
        "BB-CC": "Curl Cream",
        "BB-XYZ-2PCK": "Thing",
        "MYSTERY1": "Thing Two Pack",
        "BB-SS-2PK": "Thing",  # PK-style pack code
        "SET1": "Thing",  # trailing-digit house style (CLAY1/KIT1)
        "MYSTERY2": "Shower Essentials",  # real Based set-name shape
    }
    assert dsv.suspicious_unexploded_skus(codes, names) == [
        "BB-SS-2PK",
        "BB-XYZ-2PCK",
        "MYSTERY1",
        "MYSTERY2",
        "SET1",
    ]


def test_suspicious_unexploded_skus_ignores_real_singles() -> None:
    codes = {"44126606262501", "BB-SRS-4OZ", "CLAY1", "BB-DEO-SS-01", "BB-ACCS-SCPS"}
    names = {
        "44126606262501": "Texture Powder",
        "BB-SRS-4OZ": "Skin Revival Spray",
        "CLAY1": "Clay",
        "BB-DEO-SS-01": "Deodorant",
        "BB-ACCS-SCPS": "Scalp Scrubber",
    }
    assert dsv.suspicious_unexploded_skus(codes, names) == []


def test_nested_bundle_skus_flags_bundle_used_as_component() -> None:
    duo = BundleEntry(
        bundle_sku="DUO1",
        bundle_name="Shower Duo",
        source="shiphero",
        components_resolved=(("BB-SHMP", "Shampoo", 1), ("BB-COND", "Conditioner", 1)),
        partially_resolved=False,
    )
    deluxe = BundleEntry(
        bundle_sku="DELUXE1",
        bundle_name="Deluxe Kit",
        source="shiphero",
        components_resolved=(("DUO1", "Shower Duo", 1), ("BB-CC", "Curl Cream", 1)),
        partially_resolved=False,
    )
    registry = BundleRegistry(
        bundles=(duo, deluxe),
        bundle_skus=frozenset({"DUO1", "DELUXE1"}),
        by_bundle_sku={"DUO1": duo, "DELUXE1": deluxe},
    )
    # DUO1 is a component of DELUXE1 -> its as-sold count may include kit lines.
    assert dsv.nested_bundle_skus(registry) == {"DUO1"}
    assert dsv.nested_bundle_skus(_registry()) == set()  # no nesting -> empty


# --- ABOUT/README lines -------------------------------------------------------
def _about_lines(**overrides):
    base = dict(
        start_date=date(2026, 3, 1),
        end_date=date(2026, 3, 1),
        n_days=1,
        line_items_cap=15,
        sku_count=10,
        total_units=100,
        channel_totals=[("TIKTOK", 60), ("SHOPIFY", 40)],
        include_shops_key=["BASED"],
        excluded_shops={},
        all_active_skus=False,
        amazon_present=False,
        bundle_count=2,
        bundle_units=5,
        generated_at="2026-03-02 00:00:00",
    )
    base.update(overrides)
    return dsv.build_about_lines(**base)


def test_build_about_lines_states_bundles_have_no_rows_and_names_the_tab() -> None:
    text = "\n".join(_about_lines())
    assert "NO row of their own" in text
    assert "BUNDLE SALES" in text  # points readers at the as-sold view
    assert "do not add" in text.lower()  # warns against double-counting


def test_build_about_lines_amazon_note_only_when_amazon_present() -> None:
    with_amazon = _about_lines(amazon_present=True)
    without = _about_lines(amazon_present=False)
    assert any("FBM ONLY" in line for line in with_amazon)
    assert not any("FBM ONLY" in line for line in without)
    # The FBM caveat must come before the per-tab layout line, as today.
    fbm_i = next(i for i, line in enumerate(with_amazon) if "FBM ONLY" in line)
    tab_i = next(i for i, line in enumerate(with_amazon) if "One tab per sales channel" in line)
    assert fbm_i < tab_i


def test_build_about_lines_allowlist_note_only_when_channels_excluded() -> None:
    noisy = _about_lines(excluded_shops={"Manual Order": 3})
    quiet = _about_lines(excluded_shops={})
    assert any("strict allowlist" in line for line in noisy)
    assert not any("strict allowlist" in line for line in quiet)


# --- multi-tab / long-format writers ----------------------------------------
def test_write_xlsx_has_about_plus_one_tab_per_channel(tmp_path) -> None:
    from openpyxl import load_workbook

    rows = [("2026-03-01", "Curl Cream", 5, "BB-CC")]
    sheets = [
        (dsv._sheet_title("TIKTOK"), rows),
        (dsv._sheet_title("AMAZON"), rows),
    ]
    out = tmp_path / "v.xlsx"
    dsv.write_xlsx(sheets, out, ["about line"], include_code=True)
    wb = load_workbook(out)
    assert wb.sheetnames == ["ABOUT", "DAILY VELOCITY - TIKTOK", "DAILY VELOCITY - AMAZON"]
    ws = wb["DAILY VELOCITY - TIKTOK"]
    assert [c.value for c in ws[1]] == ["Date", "SKU", "QTY SOLD", "SKU CODE"]
    assert [c.value for c in ws[2]] == ["2026-03-01", "Curl Cream", 5, "BB-CC"]


def test_write_csv_long_format_has_channel_column(tmp_path) -> None:
    import csv

    channel_rows = [
        ("TIKTOK", [("2026-03-01", "Curl Cream", 5, "BB-CC")]),
        ("AMAZON", [("2026-03-01", "Curl Cream", 2, "BB-CC")]),
    ]
    out = tmp_path / "v.csv"
    dsv.write_csv(channel_rows, out, include_code=True)
    with open(out) as f:
        rows = list(csv.reader(f))
    assert rows[0] == ["Date", "SKU", "Channel", "QTY SOLD", "SKU CODE"]
    # Both channels present, sorted (AMAZON before TIKTOK within the same date/SKU).
    assert rows[1] == ["2026-03-01", "Curl Cream", "AMAZON", "2", "BB-CC"]
    assert rows[2] == ["2026-03-01", "Curl Cream", "TIKTOK", "5", "BB-CC"]


def test_write_csv_custom_sku_header_for_bundles_file() -> None:
    import csv
    import tempfile

    channel_rows = [("TIKTOK", [("2026-03-01", "Curly Kit", 2, "KIT1")])]
    with tempfile.TemporaryDirectory() as td:
        out = Path(td) / "b.csv"
        dsv.write_csv(channel_rows, out, include_code=True, sku_header="Bundle")
        with open(out) as f:
            rows = list(csv.reader(f))
    assert rows[0] == ["Date", "Bundle", "Channel", "QTY SOLD", "SKU CODE"]
    assert rows[1] == ["2026-03-01", "Curly Kit", "TIKTOK", "2", "KIT1"]


def test_write_xlsx_extra_long_sheets_adds_bundle_tab(tmp_path) -> None:
    from openpyxl import load_workbook

    rows = [("2026-03-01", "Curl Cream", 5, "BB-CC")]
    sheets = [(dsv._sheet_title("TIKTOK"), rows)]
    bundle_header = ["Date", "Bundle", "Channel", "QTY SOLD", "SKU CODE"]
    bundle_rows = [["2026-03-01", "Curly Kit", "TIKTOK", 2, "KIT1"]]
    out = tmp_path / "v.xlsx"
    dsv.write_xlsx(
        sheets,
        out,
        ["about line"],
        include_code=True,
        extra_long_sheets=[("BUNDLE SALES", bundle_header, bundle_rows)],
    )
    wb = load_workbook(out)
    assert wb.sheetnames == ["ABOUT", "DAILY VELOCITY - TIKTOK", "BUNDLE SALES"]
    ws = wb["BUNDLE SALES"]
    assert [c.value for c in ws[1]] == bundle_header
    assert [c.value for c in ws[2]] == ["2026-03-01", "Curly Kit", "TIKTOK", 2, "KIT1"]


def test_sheet_title_is_excel_safe() -> None:
    assert dsv._sheet_title("AMAZON") == "DAILY VELOCITY - AMAZON"
    assert len(dsv._sheet_title("A VERY LONG CHANNEL NAME THAT EXCEEDS LIMIT")) <= 31


# --- end-to-end run() wiring (faked client; no network) ----------------------
def test_run_writes_bundle_view_alongside_channel_grids(tmp_path, monkeypatch) -> None:
    """One synthetic day through run(): channel grids stay exploded-only, the
    bundle view lands in .bundles.csv + the BUNDLE SALES tab, checkpoint is v3."""
    import csv
    import json

    from openpyxl import load_workbook

    orders = [
        {
            "order_number": "1",
            "shop_name": "BASED",
            "order_date": "2026-03-01T01:00:00",
            "line_items": {
                "edges": [
                    {"node": {"sku": "KIT1", "quantity": 2}},
                    # pre-exploded component lines, as ShipHero ships them
                    {"node": {"sku": "BB-CC", "quantity": 2}},
                    {"node": {"sku": "BB-SHMP", "quantity": 4}},
                ]
            },
        },
        {
            "order_number": "2",
            "shop_name": "Manual Order",  # must be excluded end-to-end
            "order_date": "2026-03-01T02:00:00",
            "line_items": {"edges": [{"node": {"sku": "BB-CC", "quantity": 999}}]},
        },
    ]

    class FakeClient:
        def fetch_orders_window(self, since_iso, warehouse_id=None, max_pages=0, until_iso=None):
            return orders

    stock = [
        _stock("BB-CC", "Curl Cream"),
        _stock("BB-SHMP", "Shampoo"),
        _stock("KIT1", "Curly Kit", is_kit=True),
        _stock("NEWKIT", "Brand New Kit", is_kit=True),  # unregistered -> must warn
    ]
    monkeypatch.setattr(dsv, "_build_client", lambda _floor: FakeClient())
    monkeypatch.setattr(dsv, "_load_catalog", lambda _client: (stock, _registry()))

    args = dsv.build_arg_parser().parse_args(
        ["--start-date", "2026-03-01", "--end-date", "2026-03-01", "--out-dir", str(tmp_path)]
    )
    assert dsv.run(args) == 0

    slug = "daily-sales-velocity_2026-03-01_2026-03-01"
    # Channel CSV: exploded components only, Manual Order excluded, no KIT1 row.
    with open(tmp_path / f"{slug}.csv") as f:
        main_rows = list(csv.DictReader(f))
    by_sku = {}
    for r in main_rows:
        by_sku[r["SKU CODE"]] = by_sku.get(r["SKU CODE"], 0) + int(r["QTY SOLD"])
    assert by_sku == {"BB-CC": 2, "BB-SHMP": 4}
    # Bundles CSV: KIT1 as-sold on TIKTOK only.
    with open(tmp_path / f"{slug}.bundles.csv") as f:
        bundle_rows = list(csv.DictReader(f))
    assert bundle_rows[0]["Bundle"] == "Curly Kit"
    tiktok = [r for r in bundle_rows if r["Channel"] == "TIKTOK"]
    assert len(tiktok) == 1 and tiktok[0]["QTY SOLD"] == "2"
    # XLSX: BUNDLE SALES tab present with the as-sold row.
    wb = load_workbook(tmp_path / f"{slug}.xlsx")
    assert "BUNDLE SALES" in wb.sheetnames
    ws = wb["BUNDLE SALES"]
    data = [[c.value for c in row] for row in ws.iter_rows(min_row=2)]
    assert ["2026-03-01", "Curly Kit", "TIKTOK", 2, "KIT1"] in data
    # Checkpoint carries the v3 schema + the sibling bundle_days map.
    ckpt = json.loads((tmp_path / f"{slug}.checkpoint.json").read_text())
    assert ckpt["schema"] == 3
    assert ckpt["bundle_days"]["2026-03-01"]["BASED"] == {"KIT1": 2}
    # Coverage guard fired for the unregistered kit, INSIDE the README body
    # (before the Generated footer), not as trailing debris after it.
    readme = (tmp_path / f"{slug}.README.txt").read_text().splitlines()
    warn_i = next(i for i, line in enumerate(readme) if "NEWKIT" in line and "WARNING" in line)
    gen_i = next(i for i, line in enumerate(readme) if line.startswith("Generated:"))
    assert warn_i < gen_i


# --- network resilience for the long unattended pull ------------------------
def test_is_retryable_network_error_classification() -> None:
    import requests

    assert dsv._is_retryable_network_error(requests.ConnectionError("boom")) is True
    assert (
        dsv._is_retryable_network_error(
            RuntimeError("ShipHero network error after 6 retries: Failed to resolve host")
        )
        is True
    )
    # A real GraphQL error is NOT a transient network blip; must not retry forever.
    assert (
        dsv._is_retryable_network_error(
            RuntimeError("ShipHero GraphQL errors: [{'message': 'bad field'}]")
        )
        is False
    )


def test_fetch_orders_resilient_retries_then_succeeds(monkeypatch) -> None:
    monkeypatch.setattr(dsv.time, "sleep", lambda _s: None)  # no real waiting
    calls = {"n": 0}

    class FlakyClient:
        def fetch_orders_window(self, since_iso, warehouse_id=None, max_pages=0, until_iso=None):
            calls["n"] += 1
            if calls["n"] < 3:
                raise RuntimeError("ShipHero network error after 6 retries: Max retries exceeded")
            return [{"order_number": "1"}]

    out = dsv._fetch_orders_resilient(FlakyClient(), "2026-03-01", "s", "u", 10)
    assert out == [{"order_number": "1"}]
    assert calls["n"] == 3  # failed twice, succeeded on the third


def test_fetch_timeout_is_retryable() -> None:
    assert dsv._is_retryable_network_error(dsv._FetchTimeoutError("hang")) is True


def test_fetch_orders_resilient_retries_on_watchdog_timeout(monkeypatch) -> None:
    # A hung day (watchdog fires -> _FetchTimeoutError) must be retried, not fatal.
    monkeypatch.setattr(dsv.time, "sleep", lambda _s: None)
    monkeypatch.setattr(dsv, "_set_watchdog", lambda _s: None)  # don't arm a real SIGALRM in tests
    monkeypatch.setattr(dsv, "_clear_watchdog", lambda _p: None)
    calls = {"n": 0}

    class HangingClient:
        def fetch_orders_window(self, since_iso, warehouse_id=None, max_pages=0, until_iso=None):
            calls["n"] += 1
            if calls["n"] == 1:
                raise dsv._FetchTimeoutError("per-day fetch watchdog fired")
            return [{"order_number": "ok"}]

    out = dsv._fetch_orders_resilient(HangingClient(), "2026-03-01", "s", "u", 10)
    assert out == [{"order_number": "ok"}] and calls["n"] == 2


def test_fetch_orders_resilient_propagates_non_network_error(monkeypatch) -> None:
    monkeypatch.setattr(dsv, "_set_watchdog", lambda _s: None)
    monkeypatch.setattr(dsv, "_clear_watchdog", lambda _p: None)
    monkeypatch.setattr(dsv.time, "sleep", lambda _s: None)

    class BadQueryClient:
        def fetch_orders_window(self, since_iso, warehouse_id=None, max_pages=0, until_iso=None):
            raise RuntimeError("ShipHero GraphQL errors: [{'message': 'bad field'}]")

    try:
        dsv._fetch_orders_resilient(BadQueryClient(), "2026-03-01", "s", "u", 10)
        raise AssertionError("expected the GraphQL error to propagate")
    except RuntimeError as e:
        assert "GraphQL" in str(e)


# --- additive until_iso bound on the reused fetcher --------------------------
def test_fetch_orders_window_until_iso_is_upper_bound(monkeypatch) -> None:
    client = ShipHeroClient(token="x")
    captured: dict = {}

    def fake_execute(query, variables=None, retries=6):
        captured.update(variables or {})
        return {"data": {"orders": {"data": {"edges": [], "pageInfo": {"hasNextPage": False}}}}}

    monkeypatch.setattr(client, "_execute", fake_execute)
    client.fetch_orders_window(since_iso="2026-03-01T00:00:00", until_iso="2026-03-02T00:00:00")
    assert captured["since"] == "2026-03-01T00:00:00"
    assert captured["until"] == "2026-03-02T00:00:00"


def test_fetch_orders_window_defaults_until_to_now(monkeypatch) -> None:
    client = ShipHeroClient(token="x")
    captured: dict = {}

    def fake_execute(query, variables=None, retries=6):
        captured.update(variables or {})
        return {"data": {"orders": {"data": {"edges": [], "pageInfo": {"hasNextPage": False}}}}}

    monkeypatch.setattr(client, "_execute", fake_execute)
    client.fetch_orders_window(since_iso="2026-03-01T00:00:00")
    # No explicit upper bound -> a concrete 'YYYY-MM-DDTHH:MM:SS' now-timestamp.
    assert len(captured["until"]) == 19
    assert captured["until"] > captured["since"]
